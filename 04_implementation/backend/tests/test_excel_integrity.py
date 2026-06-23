from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import openpyxl

from app.core.paths import WORKBOOK_PATH
from app.excel.apu_repository import ApuRepository
from app.excel.budget_repository import BudgetRepository
from app.excel.formulas import collect_workbook_formulas, compare_formulas
from app.excel.materials_repository import MaterialsRepository
from app.excel.workbook import ExcelWorkbook
from app.schemas.apus import ApuCreate, ApuDetail
from app.schemas.budget import BudgetApuCreate
from app.services.materials_service import MaterialsService
from app.schemas.materials import MaterialCreate


def test_can_read_materials_and_apus():
    excel = ExcelWorkbook(WORKBOOK_PATH, WORKBOOK_PATH.parent / "backups")

    materials = MaterialsRepository(excel).list_materials()
    apus = ApuRepository(excel).list_apus()

    assert len(materials) > 300
    assert len(apus) > 1
    assert any(material.codigo == "ARAN-P1/2" for material in materials)


def test_add_material_on_copy_preserves_existing_formulas():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        excel = ExcelWorkbook(temp_path, backup_dir)
        before_book = excel.load()
        expected_row = before_book["L-MATERIALES"].max_row + 1
        before_formulas = collect_workbook_formulas(before_book)
        before_book.close()

        result = MaterialsRepository(excel).add_material(
            MaterialCreate(
                codigo="TEST-MAT-001",
                descripcion="Material de prueba",
                marca="TEST",
                und="UN",
                factor=0.2,
                valor_costo_incluido_iva=1000,
            )
        )

        after_book = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            after_formulas = collect_workbook_formulas(after_book)
            report = compare_formulas(before_formulas, after_formulas)
            material_sheet = after_book["L-MATERIALES"]

            assert result["formula_integrity_ok"] is True
            assert result["row_number"] == expected_row
            assert material_sheet[f"B{expected_row}"].value == "TEST-MAT-001"
            assert material_sheet[f"H{expected_row}"].value == f"=G{expected_row}/(100%-F{expected_row})"
            assert material_sheet[f"J{expected_row}"].value == f"=H{expected_row}*I{expected_row}"
            assert not report.changed
            assert not report.removed
            assert (backup_dir / temp_path.name).exists() is False
            assert any(backup_dir.glob("PRESUPUESTO-TULUA2_*.xlsx"))
        finally:
            after_book.close()


def test_material_service_suggests_unique_code():
    class FakeRepository:
        def exists(self, codigo: str) -> bool:
            return codigo in {"TUBO-PVC-3/4", "TUBO-PVC-3/4-2"}

    service = MaterialsService(repository=FakeRepository())

    assert service.suggest_code("Tubo PVC 3/4") == "TUBO-PVC-3/4-3"


def test_get_apu_draft_from_existing_row_returns_materials_only():
    excel = ExcelWorkbook(WORKBOOK_PATH, WORKBOOK_PATH.parent / "backups")
    repository = ApuRepository(excel)
    base_apu = repository.search("711")[0]

    draft = repository.get_draft_from_row(base_apu.row_number)

    assert draft.base_codigo_apu == base_apu.codigo_apu
    assert draft.descripcion_apu == base_apu.descripcion_apu
    assert draft.unidad == base_apu.unidad
    assert len(draft.detalles) >= 1
    assert all(detail.tipo_item == "MATERIAL" for detail in draft.detalles)
    assert not {"E&H", "GLV", "T", "H/C"}.intersection(
        {detail.codigo_item for detail in draft.detalles}
    )
    assert all(detail.descripcion for detail in draft.detalles)
    assert all(detail.cantidad for detail in draft.detalles)


def test_get_apu_draft_allows_formula_quantities():
    excel = ExcelWorkbook(WORKBOOK_PATH, WORKBOOK_PATH.parent / "backups")
    repository = ApuRepository(excel)

    draft = repository.get_draft_from_row(941)
    formula_detail = next(
        detail for detail in draft.detalles if detail.codigo_item == "BORN-TER 350"
    )

    assert formula_detail.source_row_number == 945
    assert formula_detail.cantidad == "0"
    assert formula_detail.cantidad_formula == "=IFERROR(((F943*2*K945)/J944),0)"


def test_create_apu_translates_formula_quantities_from_base_rows():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        excel = ExcelWorkbook(temp_path, backup_dir)
        repository = ApuRepository(excel)
        draft = repository.get_draft_from_row(941)
        selected_details = [
            detail
            for detail in draft.detalles
            if detail.codigo_item in {"THHN 350 KCMIL", "THHN 2/0 AWG", "BORN-TER 350"}
        ]

        result = repository.create_apu(
            ApuCreate(
                codigo_apu="TEST-FORMULA-APU",
                descripcion_apu="APU con formula base",
                unidad="ML",
                detalles=[
                    ApuDetail(
                        tipo_item="MATERIAL",
                        codigo_item=detail.codigo_item,
                        cantidad=detail.cantidad,
                        source_row_number=detail.source_row_number,
                        cantidad_formula=detail.cantidad_formula,
                    )
                    for detail in selected_details
                ],
            )
        )

        workbook = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            sheet = workbook["APU'S"]
            header_row = result["header_row"]
            formula_row = header_row + 4

            assert result["formula_integrity_ok"] is True
            assert sheet.cell(row=formula_row, column=2).value == "BORN-TER 350"
            assert sheet.cell(row=formula_row, column=6).value == (
                f"=IFERROR(((F{formula_row - 2}*2*K{formula_row})/J{formula_row - 1}),0)"
            )
        finally:
            workbook.close()


def test_create_apu_adds_standard_formula_rows():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        excel = ExcelWorkbook(temp_path, backup_dir)
        result = ApuRepository(excel).create_apu(
            ApuCreate(
                codigo_apu="TEST-APU-001",
                descripcion_apu="APU de prueba",
                unidad="UN",
                detalles=[
                    ApuDetail(
                        tipo_item="MATERIAL",
                        codigo_item="ARAN-P1/2",
                        cantidad=2,
                    )
                ],
            )
        )

        workbook = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            sheet = workbook["APU'S"]
            header_row = result["header_row"]
            material_row = header_row + 2
            standard_start = material_row + 1
            standard_end = standard_start + 3
            blank_row = standard_end + 1

            assert result["formula_integrity_ok"] is True
            assert result["detail_rows"] == 6
            assert sheet.cell(row=header_row, column=8).value == (
                f"=SUM(H{material_row}:H{standard_end})"
            )
            assert sheet.cell(row=header_row, column=9).value == (
                f"=+SUM(H{standard_start}:H{standard_end})"
            )
            assert sheet.cell(row=material_row, column=2).value == "ARAN-P1/2"
            assert sheet.cell(row=material_row, column=8).value == (
                f"=F{material_row}*G{material_row}"
            )
            assert [
                sheet.cell(row=row, column=2).value
                for row in range(standard_start, standard_end + 1)
            ] == ["E&H", "GLV", "T", "H/C"]
            assert all(
                isinstance(sheet.cell(row=row, column=8).value, str)
                and sheet.cell(row=row, column=8).value.startswith("=")
                for row in range(standard_start, standard_end + 1)
            )
            assert sheet.cell(row=header_row, column=1).value == (
                f'=+IF(B{blank_row}="",B{header_row},A{blank_row})'
            )
            assert sheet.cell(row=blank_row, column=1).value == (
                f'=+IF(B{blank_row - 1}="",B{blank_row},A{blank_row - 1})'
            )
            assert sheet.cell(row=blank_row, column=2).value is None
            assert all(
                isinstance(sheet.cell(row=row, column=1).value, str)
                and sheet.cell(row=row, column=1).value.startswith("=")
                for row in range(header_row, blank_row + 1)
            )
        finally:
            workbook.close()


def test_add_apu_to_budget_creates_presupuesto_sheet_with_formulas():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        workbook = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            if "PRESUPUESTO" in workbook.sheetnames:
                workbook.remove(workbook["PRESUPUESTO"])
                workbook.save(temp_path)
        finally:
            workbook.close()

        excel = ExcelWorkbook(temp_path, backup_dir)
        result = BudgetRepository(excel).add_apu(
            BudgetApuCreate(codigo_apu="711", cantidad=2)
        )

        workbook = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            assert "PRESUPUESTO" in workbook.sheetnames
            sheet = workbook["PRESUPUESTO"]
            row = result["row_number"]

            assert result["formula_integrity_ok"] is True
            assert row >= 6
            assert sheet.cell(row=row, column=2).value >= 1
            assert sheet.cell(row=row, column=3).value == 711
            assert sheet.cell(row=row, column=4).value == f"=+VLOOKUP(C{row},APU_S,2,0)"
            assert sheet.cell(row=row, column=5).value == f"=+VLOOKUP(C{row},APU_S,4,0)"
            assert sheet.cell(row=row, column=6).value == 2
            assert sheet.cell(row=row, column=7).value == f"=+I{row}-H{row}"
            assert sheet.cell(row=row, column=8).value == f"=+VLOOKUP(C{row},APU_S,8,0)"
            assert sheet.cell(row=row, column=9).value == f"=+VLOOKUP(C{row},APU_S,7,0)"
            assert sheet.cell(row=row, column=10).value == f"=ROUND(F{row}*I{row},0)"
            assert sheet.cell(row=row, column=3).border.left.style is not None
        finally:
            workbook.close()


def test_save_budget_creates_new_sheet_with_template_columns_and_formulas():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        excel = ExcelWorkbook(temp_path, backup_dir)
        result = BudgetRepository(excel).save_budget(
            [
                BudgetApuCreate(codigo_apu="711", cantidad=2),
                BudgetApuCreate(codigo_apu="550", cantidad=1),
            ]
        )

        workbook = openpyxl.load_workbook(temp_path, data_only=False, read_only=False)
        try:
            template = workbook["PRESUPUESTO UNIFORCE"]
            sheet = workbook[result["sheet_name"]]
            summary_row = 8

            assert result["formula_integrity_ok"] is True
            assert result["rows"] == 2
            assert sheet.max_column >= template.max_column
            merged_ranges = {str(item) for item in sheet.merged_cells.ranges}
            assert "B1:J1" in merged_ranges
            assert "B2:J2" in merged_ranges
            assert "B3:J3" in merged_ranges
            assert "B10:D16" in merged_ranges
            assert sheet.cell(row=5, column=4).value == template.cell(row=5, column=4).value
            assert sheet.cell(row=6, column=2).value == 1
            assert sheet.cell(row=7, column=2).value == 2
            assert sheet.cell(row=6, column=3).value == 711
            assert sheet.cell(row=7, column=3).value == 550
            assert sheet.cell(row=6, column=4).value == "=+VLOOKUP(C6,APU_S,2,0)"
            assert sheet.cell(row=7, column=4).value == "=+VLOOKUP(C7,APU_S,2,0)"
            assert sheet.cell(row=6, column=10).value == "=ROUND(F6*I6,0)"
            assert sheet.cell(row=7, column=10).value == "=ROUND(F7*I7,0)"
            assert sheet.cell(row=summary_row, column=6).value == (
                '=CONCATENATE("SUBTOTAL ",D5)'
            )
            assert sheet.cell(row=summary_row, column=7).value == (
                "=+SUMPRODUCT(G6:G7,F6:F7)"
            )
            assert sheet.cell(row=summary_row, column=8).value == (
                "=+SUMPRODUCT(H6:H7,F6:F7)"
            )
            assert sheet.cell(row=summary_row, column=10).value == "=SUM(J6:L7)"
            assert sheet.cell(row=summary_row + 2, column=2).value == "UNIFORCE SAS"
            assert sheet.cell(row=summary_row + 2, column=10).value == (
                "=#REF!+J8+#REF!+#REF!+#REF!+#REF!+#REF!"
            )
            assert sheet.max_row == 48
            assert sheet.cell(row=6, column=3).border.left.style is not None
        finally:
            workbook.close()


def test_export_filtered_budget_creates_workbook_with_only_used_apus_and_materials():
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / WORKBOOK_PATH.name
        backup_dir = Path(temp_dir) / "backups"
        export_dir = Path(temp_dir) / "exports"
        shutil.copy2(WORKBOOK_PATH, temp_path)

        excel = ExcelWorkbook(temp_path, backup_dir)
        result = BudgetRepository(excel, exports_dir=export_dir).export_filtered_budget(
            [
                BudgetApuCreate(codigo_apu="711", cantidad=2),
                BudgetApuCreate(codigo_apu="550", cantidad=1),
                BudgetApuCreate(codigo_apu="711", cantidad=3),
            ],
            export_code="PRUEBA-001",
        )

        export_path = Path(result["file_path"])
        assert export_path.exists()
        assert export_path.parent == export_dir
        assert result["file_name"].startswith("PRUEBA-001_")
        assert result["rows"] == 3
        assert result["apus"] == 2

        workbook = openpyxl.load_workbook(export_path, data_only=False, read_only=False)
        try:
            assert "PRESUPUESTO" in workbook.sheetnames
            assert "PRESUPUESTO 2" not in workbook.sheetnames
            assert "PRESUPUESTO UNIFORCE" in workbook.sheetnames
            assert "MANO DE OBRA" in workbook.sheetnames
            assert "APU'S" in workbook.sheetnames
            assert "L-MATERIALES" in workbook.sheetnames

            budget_sheet = workbook["PRESUPUESTO"]
            assert budget_sheet.cell(row=6, column=3).value == 711
            assert budget_sheet.cell(row=7, column=3).value == 550
            assert budget_sheet.cell(row=8, column=3).value == 711

            apu_sheet = workbook["APU'S"]
            apu_codes = [
                str(apu_sheet.cell(row=row, column=2).value).strip()
                for row in range(1, apu_sheet.max_row + 1)
                if apu_sheet.cell(row=row, column=7).value in {"COSTO ITEM", "COSTOITEM"}
            ]
            assert apu_codes == ["711", "550"]

            used_materials: set[str] = set()
            for row in range(1, apu_sheet.max_row + 1):
                code = apu_sheet.cell(row=row, column=2).value
                if code in (None, ""):
                    continue
                normalized = str(code).strip()
                if normalized in {"711", "550", "CODIGO", "E&H", "GLV", "T", "H/C"}:
                    continue
                used_materials.add(normalized)

            materials_sheet = workbook["L-MATERIALES"]
            material_codes = {
                str(materials_sheet.cell(row=row, column=2).value).strip()
                for row in range(5, materials_sheet.max_row + 1)
                if materials_sheet.cell(row=row, column=2).value not in (None, "")
            }
            assert material_codes == used_materials
            assert result["materials"] == len(used_materials)
        finally:
            workbook.close()
