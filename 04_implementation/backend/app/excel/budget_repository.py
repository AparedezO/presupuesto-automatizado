from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl.formula.translate import Translator
from app.excel.formulas import collect_workbook_formulas
from app.excel.workbook import ExcelWorkbook, find_last_data_row
from app.schemas.budget import BudgetApuCreate, BudgetItemRead
from openpyxl.utils import get_column_letter, range_boundaries


BUDGET_SHEET = "PRESUPUESTO"
TEMPLATE_BUDGET_SHEET = "PRESUPUESTO UNIFORCE"
APU_SHEET = "APU'S"
MATERIALS_SHEET = "L-MATERIALES"
HEADER_ROW = 4
SECTION_ROW = 5
FIRST_DATA_ROW = 6
LAST_DATA_ROW = 31
SUMMARY_START_ROW = 32
CODE_COL = 3
SUMMARY_END_ROW = 72
CELL_REFERENCE_PATTERN = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})(\$?)(\d+)\b")
GENERATED_BUDGET_PATTERN = re.compile(r"^PRESUPUESTO(?: \d+)?$")
STANDARD_APU_CODES = {"E&H", "GLV", "T", "H/C"}
MATERIAL_FIRST_DATA_ROW = 5
MATERIAL_CODE_COL = 2


def _decimal_or_one(value) -> Decimal:
    if value in (None, ""):
        return Decimal("1")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("1")


def _copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _copy_row_format(
    source_sheet,
    target_sheet,
    source_row: int,
    target_row: int,
    max_column: int | None = None,
) -> None:
    source_dimension = source_sheet.row_dimensions[source_row]
    target_dimension = target_sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    max_column = max_column or source_sheet.max_column
    for column in range(1, max_column + 1):
        _copy_cell_format(
            source_sheet.cell(row=source_row, column=column),
            target_sheet.cell(row=target_row, column=column),
        )


def _shift_summary_formula(formula: str, row_offset: int) -> str:
    if row_offset == 0:
        return formula

    def replace_reference(match: re.Match[str]) -> str:
        if match.start() > 0 and formula[match.start() - 1] == "!":
            return match.group(0)

        row_number = int(match.group(3))
        if row_number < SUMMARY_START_ROW:
            return match.group(0)

        return f"{match.group(1)}{match.group(2)}{row_number + row_offset}"

    return CELL_REFERENCE_PATTERN.sub(replace_reference, formula)


def _copy_column_dimensions(source_sheet, target_sheet) -> None:
    for column in range(1, source_sheet.max_column + 1):
        letter = get_column_letter(column)
        source_dimension = source_sheet.column_dimensions[letter]
        target_dimension = target_sheet.column_dimensions[letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed


def _copy_template_row(
    source_sheet,
    target_sheet,
    source_row: int,
    target_row: int,
    row_offset: int | None = None,
) -> None:
    source_dimension = source_sheet.row_dimensions[source_row]
    target_dimension = target_sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, source_sheet.max_column + 1):
        source = source_sheet.cell(row=source_row, column=column)
        target = target_sheet.cell(row=target_row, column=column)
        _copy_cell_format(source, target)

        if (
            row_offset is not None
            and isinstance(source.value, str)
            and source.value.startswith("=")
        ):
            target.value = _shift_summary_formula(source.value, row_offset)
        else:
            target.value = source.value


def _copy_merged_ranges(
    source_sheet,
    target_sheet,
    source_start_row: int,
    source_end_row: int,
    row_offset: int,
) -> None:
    for merged_range in source_sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row < source_start_row or max_row > source_end_row:
            continue

        target_range = (
            f"{get_column_letter(min_col)}{min_row + row_offset}:"
            f"{get_column_letter(max_col)}{max_row + row_offset}"
        )
        target_sheet.merge_cells(target_range)


def _clear_sheet_cells(sheet) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))

    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None


def _safe_sheet_name(workbook, base_name: str) -> str:
    clean_name = (base_name or BUDGET_SHEET).strip()[:31] or BUDGET_SHEET
    if clean_name not in workbook.sheetnames:
        return clean_name

    suffix = 2
    while True:
        suffix_text = f" {suffix}"
        candidate = f"{clean_name[:31 - len(suffix_text)]}{suffix_text}"
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def _safe_file_code(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", value.strip()).strip("-_")
    return safe[:80] or "PRESUPUESTO"


def _normalized_code(value) -> str:
    return str(value).strip().upper()


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _copy_row_between_sheets(
    source_sheet,
    target_sheet,
    source_row: int,
    target_row: int,
    max_column: int | None = None,
) -> set[str]:
    max_column = max_column or source_sheet.max_column
    added_formulas: set[str] = set()
    source_dimension = source_sheet.row_dimensions[source_row]
    target_dimension = target_sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, max_column + 1):
        source = source_sheet.cell(row=source_row, column=column)
        target = target_sheet.cell(row=target_row, column=column)
        _copy_cell_format(source, target)

        if _is_formula(source.value):
            try:
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value
            added_formulas.add(f"{target_sheet.title}!{target.coordinate}")
        else:
            target.value = source.value

    return added_formulas


def _delete_rows_from(sheet, start_row: int) -> None:
    if sheet.max_row >= start_row:
        sheet.delete_rows(start_row, sheet.max_row - start_row + 1)


class BudgetRepository:
    def __init__(
        self,
        excel: ExcelWorkbook | None = None,
        exports_dir: Path | None = None,
    ) -> None:
        self.excel = excel or ExcelWorkbook()
        from app.core.config import settings

        self.exports_dir = exports_dir or settings.exports_dir

    def _apu_lookup(self, workbook) -> dict[str, dict[str, str]]:
        sheet = workbook[APU_SHEET]
        lookup: dict[str, dict[str, str]] = {}
        for row in range(1, sheet.max_row + 1):
            marker = sheet.cell(row=row, column=7).value
            code = sheet.cell(row=row, column=2).value
            if marker not in {"COSTO ITEM", "COSTOITEM"} or code in (None, ""):
                continue
            normalized_code = str(code).strip()
            lookup[normalized_code.upper()] = {
                "raw_codigo_apu": code,
                "codigo_apu": normalized_code,
                "descripcion_apu": str(sheet.cell(row=row, column=3).value or "").strip(),
                "unidad": str(sheet.cell(row=row, column=5).value or "").strip(),
                "costo_total_formula": str(sheet.cell(row=row, column=8).value or ""),
            }
        return lookup

    def _ensure_sheet(self, workbook):
        if BUDGET_SHEET in workbook.sheetnames:
            return workbook[BUDGET_SHEET]

        template = workbook[TEMPLATE_BUDGET_SHEET]
        sheet = workbook.create_sheet(BUDGET_SHEET)

        for column in range(1, 11):
            letter = template.cell(row=HEADER_ROW, column=column).column_letter
            sheet.column_dimensions[letter].width = template.column_dimensions[letter].width

        for row in range(1, HEADER_ROW + 1):
            sheet.row_dimensions[row].height = template.row_dimensions[row].height
            for column in range(1, 11):
                source = template.cell(row=row, column=column)
                target = sheet.cell(row=row, column=column)
                target.value = source.value
                _copy_cell_format(source, target)

        sheet["B2"].value = "PRESUPUESTO"
        return sheet

    def _create_budget_sheet(self, workbook, sheet_name: str, item_count: int):
        template = workbook[TEMPLATE_BUDGET_SHEET]
        target_name = _safe_sheet_name(workbook, sheet_name)
        sheet = workbook.copy_worksheet(template)
        sheet.title = target_name
        summary_row = FIRST_DATA_ROW + item_count
        summary_offset = summary_row - SUMMARY_START_ROW
        final_row = summary_row + (SUMMARY_END_ROW - SUMMARY_START_ROW)

        _clear_sheet_cells(sheet)
        _copy_column_dimensions(template, sheet)

        for row in range(1, SECTION_ROW + 1):
            _copy_template_row(template, sheet, row, row)

        for offset, source_row in enumerate(range(SUMMARY_START_ROW, SUMMARY_END_ROW + 1)):
            _copy_template_row(
                template,
                sheet,
                source_row=source_row,
                target_row=summary_row + offset,
                row_offset=summary_offset,
            )

        _copy_merged_ranges(template, sheet, 1, SECTION_ROW, 0)
        _copy_merged_ranges(
            template,
            sheet,
            SUMMARY_START_ROW,
            SUMMARY_END_ROW,
            summary_offset,
        )

        if sheet.max_row > final_row:
            sheet.delete_rows(final_row + 1, sheet.max_row - final_row)

        return sheet, summary_row

    def _write_budget_row(
        self,
        sheet,
        template,
        row_number: int,
        item_number: int,
        payload: BudgetApuCreate,
        raw_codigo_apu,
    ) -> None:
        _copy_row_format(template, sheet, row_number, row_number, template.max_column)

        for column in range(1, template.max_column + 1):
            source = template.cell(row=row_number, column=column)
            target = sheet.cell(row=row_number, column=column)
            if isinstance(source.value, str) and source.value.startswith("="):
                target.value = source.value
            else:
                target.value = source.value

        sheet.cell(row=row_number, column=1).value = None
        sheet.cell(row=row_number, column=2).value = item_number
        sheet.cell(row=row_number, column=3).value = raw_codigo_apu
        sheet.cell(row=row_number, column=6).value = float(payload.cantidad)

    def _populate_budget_sheet(
        self,
        workbook,
        items: list[BudgetApuCreate],
        sheet_name: str,
    ):
        apus = self._apu_lookup(workbook)
        if len(items) > (LAST_DATA_ROW - FIRST_DATA_ROW + 1):
            raise ValueError(
                f"El modelo PRESUPUESTO UNIFORCE permite hasta {LAST_DATA_ROW - FIRST_DATA_ROW + 1} APUs antes de los totales."
            )
        for item in items:
            if item.codigo_apu.strip().upper() not in apus:
                raise ValueError(f"APU code does not exist: {item.codigo_apu}")

        sheet, summary_row = self._create_budget_sheet(workbook, sheet_name, len(items))
        template = workbook[TEMPLATE_BUDGET_SHEET]

        for index, item in enumerate(items, start=1):
            apu = apus[item.codigo_apu.strip().upper()]
            self._write_budget_row(
                sheet=sheet,
                template=template,
                row_number=FIRST_DATA_ROW + index - 1,
                item_number=index,
                payload=item,
                raw_codigo_apu=apu["raw_codigo_apu"],
            )

        last_item_row = summary_row - 1
        sheet.cell(row=summary_row, column=7).value = (
            f"=+SUMPRODUCT(G{FIRST_DATA_ROW}:G{last_item_row},"
            f"F{FIRST_DATA_ROW}:F{last_item_row})"
        )
        sheet.cell(row=summary_row, column=8).value = (
            f"=+SUMPRODUCT(H{FIRST_DATA_ROW}:H{last_item_row},"
            f"F{FIRST_DATA_ROW}:F{last_item_row})"
        )
        sheet.cell(row=summary_row, column=10).value = (
            f"=SUM(J{FIRST_DATA_ROW}:L{last_item_row})"
        )

        return sheet

    def _apu_blocks(self, sheet) -> dict[str, tuple[int, int]]:
        header_rows: list[int] = []
        for row in range(1, sheet.max_row + 1):
            marker = sheet.cell(row=row, column=7).value
            code = sheet.cell(row=row, column=2).value
            if marker in {"COSTO ITEM", "COSTOITEM"} and code not in (None, ""):
                header_rows.append(row)

        blocks: dict[str, tuple[int, int]] = {}
        for index, header_row in enumerate(header_rows):
            next_header = (
                header_rows[index + 1]
                if index + 1 < len(header_rows)
                else sheet.max_row + 1
            )
            code = _normalized_code(sheet.cell(row=header_row, column=2).value)
            blocks[code] = (header_row, next_header - 1)

        return blocks

    def _material_row_lookup(self, sheet) -> dict[str, int]:
        lookup: dict[str, int] = {}
        for row in range(MATERIAL_FIRST_DATA_ROW, sheet.max_row + 1):
            code = sheet.cell(row=row, column=MATERIAL_CODE_COL).value
            if code in (None, ""):
                continue
            lookup[_normalized_code(code)] = row
        return lookup

    def _material_codes_for_blocks(
        self,
        apu_sheet,
        blocks: list[tuple[int, int]],
    ) -> list[str]:
        seen: set[str] = set()
        codes: list[str] = []
        for start_row, end_row in blocks:
            for row in range(start_row + 2, end_row + 1):
                code = apu_sheet.cell(row=row, column=2).value
                if code in (None, ""):
                    break
                normalized_code = _normalized_code(code)
                if normalized_code in STANDARD_APU_CODES:
                    break
                if normalized_code not in seen:
                    seen.add(normalized_code)
                    codes.append(normalized_code)
        return codes

    def _filter_apu_sheet(
        self,
        source_workbook,
        target_workbook,
        budget_codes: list[str],
    ) -> tuple[int, list[str]]:
        source_sheet = source_workbook[APU_SHEET]
        target_sheet = target_workbook[APU_SHEET]
        source_blocks = self._apu_blocks(source_sheet)
        selected_blocks: list[tuple[int, int]] = []

        for code in budget_codes:
            block = source_blocks.get(code)
            if block is None:
                raise ValueError(f"APU code does not exist: {code}")
            selected_blocks.append(block)

        _delete_rows_from(target_sheet, 2)
        target_row = 2
        for start_row, end_row in selected_blocks:
            for source_row in range(start_row, end_row + 1):
                _copy_row_between_sheets(
                    source_sheet,
                    target_sheet,
                    source_row,
                    target_row,
                    source_sheet.max_column,
                )
                target_row += 1

        used_material_codes = self._material_codes_for_blocks(source_sheet, selected_blocks)
        return len(selected_blocks), used_material_codes

    def _filter_materials_sheet(
        self,
        source_workbook,
        target_workbook,
        material_codes: list[str],
    ) -> int:
        source_sheet = source_workbook[MATERIALS_SHEET]
        target_sheet = target_workbook[MATERIALS_SHEET]
        material_rows = self._material_row_lookup(source_sheet)
        missing = [code for code in material_codes if code not in material_rows]
        if missing:
            raise ValueError(
                "No se encontraron materiales usados por los APUs: "
                + ", ".join(missing[:10])
            )

        _delete_rows_from(target_sheet, MATERIAL_FIRST_DATA_ROW)
        target_row = MATERIAL_FIRST_DATA_ROW
        copied = 0
        selected = set(material_codes)
        for source_row in range(MATERIAL_FIRST_DATA_ROW, source_sheet.max_row + 1):
            code = source_sheet.cell(row=source_row, column=MATERIAL_CODE_COL).value
            if code in (None, "") or _normalized_code(code) not in selected:
                continue
            _copy_row_between_sheets(
                source_sheet,
                target_sheet,
                source_row,
                target_row,
                source_sheet.max_column,
            )
            target_row += 1
            copied += 1

        last_row = max(MATERIAL_FIRST_DATA_ROW, target_row - 1)
        if copied:
            target_sheet["J3"].value = f"=SUBTOTAL(9,J5:J{last_row})"
            target_sheet["K3"].value = f"=+SUMPRODUCT(G5:G{last_row},I5:I{last_row})"

        return copied

    def _remove_generated_budget_sheets(self, workbook) -> None:
        for sheet_name in list(workbook.sheetnames):
            if (
                sheet_name not in {TEMPLATE_BUDGET_SHEET, "PRESUPUESTO PRESENTACION"}
                and GENERATED_BUDGET_PATTERN.match(sheet_name)
            ):
                workbook.remove(workbook[sheet_name])

    def export_filtered_budget(
        self,
        items: list[BudgetApuCreate],
        export_code: str,
        sheet_name: str = BUDGET_SHEET,
    ) -> dict:
        if not items:
            raise ValueError("Agrega al menos un APU antes de exportar.")

        source_workbook = self.excel.load()
        workbook = self.excel.load()
        try:
            self._remove_generated_budget_sheets(workbook)
            sheet = self._populate_budget_sheet(workbook, items, sheet_name)

            budget_codes: list[str] = []
            seen_budget_codes: set[str] = set()
            for item in items:
                normalized_code = _normalized_code(item.codigo_apu)
                if normalized_code not in seen_budget_codes:
                    seen_budget_codes.add(normalized_code)
                    budget_codes.append(normalized_code)
            apu_count, material_codes = self._filter_apu_sheet(
                source_workbook,
                workbook,
                budget_codes,
            )
            material_count = self._filter_materials_sheet(
                source_workbook,
                workbook,
                material_codes,
            )

            self.exports_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_code = _safe_file_code(export_code)
            file_name = f"{safe_code}_{timestamp}.xlsx"
            file_path = self.exports_dir / file_name

            if hasattr(workbook, "calculation"):
                workbook.calculation.calcMode = "auto"
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True

            workbook.save(file_path)
            workbook.close()
            source_workbook.close()

            return {
                "file_name": file_name,
                "file_path": str(file_path),
                "sheet_name": sheet.title,
                "rows": len(items),
                "apus": apu_count,
                "materials": material_count,
            }
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            try:
                source_workbook.close()
            except Exception:
                pass

    def list_items(self) -> list[BudgetItemRead]:
        workbook = self.excel.load()
        try:
            if BUDGET_SHEET not in workbook.sheetnames:
                return []
            sheet = workbook[BUDGET_SHEET]
            apus = self._apu_lookup(workbook)
            last_row = find_last_data_row(sheet, CODE_COL, FIRST_DATA_ROW)
            items: list[BudgetItemRead] = []

            for row in range(FIRST_DATA_ROW, last_row + 1):
                code = sheet.cell(row=row, column=CODE_COL).value
                if code in (None, ""):
                    continue
                normalized_code = str(code).strip()
                apu = apus.get(normalized_code.upper(), {})
                items.append(
                    BudgetItemRead(
                        row_number=row,
                        item=int(sheet.cell(row=row, column=2).value or len(items) + 1),
                        codigo_apu=normalized_code,
                        descripcion_apu=apu.get("descripcion_apu", ""),
                        unidad=apu.get("unidad", ""),
                        cantidad=_decimal_or_one(sheet.cell(row=row, column=6).value),
                        material_formula=sheet.cell(row=row, column=7).value,
                        mano_obra_formula=sheet.cell(row=row, column=8).value,
                        valor_unitario_formula=sheet.cell(row=row, column=9).value,
                        valor_total_formula=sheet.cell(row=row, column=10).value,
                    )
                )

            return items
        finally:
            workbook.close()

    def add_apu(self, payload: BudgetApuCreate) -> dict:
        workbook = self.excel.load()
        try:
            before_formulas = collect_workbook_formulas(workbook)
            apus = self._apu_lookup(workbook)
            normalized_code = payload.codigo_apu.strip()
            apu = apus.get(normalized_code.upper())
            if apu is None:
                raise ValueError(f"APU code does not exist: {payload.codigo_apu}")

            sheet = self._ensure_sheet(workbook)
            template = workbook[TEMPLATE_BUDGET_SHEET]
            last_row = find_last_data_row(sheet, CODE_COL, FIRST_DATA_ROW)
            target_row = max(FIRST_DATA_ROW, last_row + 1)
            item_number = target_row - FIRST_DATA_ROW + 1

            _copy_row_format(template, sheet, 6, target_row)

            sheet.cell(row=target_row, column=1).value = None
            sheet.cell(row=target_row, column=2).value = item_number
            sheet.cell(row=target_row, column=3).value = apu["raw_codigo_apu"]
            sheet.cell(row=target_row, column=4).value = f"=+VLOOKUP(C{target_row},APU_S,2,0)"
            sheet.cell(row=target_row, column=5).value = f"=+VLOOKUP(C{target_row},APU_S,4,0)"
            sheet.cell(row=target_row, column=6).value = float(payload.cantidad)
            sheet.cell(row=target_row, column=7).value = f"=+I{target_row}-H{target_row}"
            sheet.cell(row=target_row, column=8).value = f"=+VLOOKUP(C{target_row},APU_S,8,0)"
            sheet.cell(row=target_row, column=9).value = f"=+VLOOKUP(C{target_row},APU_S,7,0)"
            sheet.cell(row=target_row, column=10).value = f"=ROUND(F{target_row}*I{target_row},0)"

            backup_path, report = self.excel.save_preserving_formulas(
                workbook,
                before_formulas,
            )

            item = BudgetItemRead(
                row_number=target_row,
                item=item_number,
                codigo_apu=normalized_code,
                descripcion_apu=apu["descripcion_apu"],
                unidad=apu["unidad"],
                cantidad=payload.cantidad,
                material_formula=sheet.cell(row=target_row, column=7).value,
                mano_obra_formula=sheet.cell(row=target_row, column=8).value,
                valor_unitario_formula=sheet.cell(row=target_row, column=9).value,
                valor_total_formula=sheet.cell(row=target_row, column=10).value,
            )

            return {
                "item": item,
                "row_number": target_row,
                "backup_path": str(backup_path),
                "formula_integrity_ok": report.ok,
            }
        finally:
            workbook.close()

    def save_budget(self, items: list[BudgetApuCreate], sheet_name: str = BUDGET_SHEET) -> dict:
        workbook = self.excel.load()
        try:
            before_formulas = collect_workbook_formulas(workbook)
            sheet = self._populate_budget_sheet(workbook, items, sheet_name)

            backup_path, report = self.excel.save_preserving_formulas(
                workbook,
                before_formulas,
            )

            return {
                "sheet_name": sheet.title,
                "rows": len(items),
                "backup_path": str(backup_path),
                "formula_integrity_ok": report.ok,
            }
        finally:
            workbook.close()
