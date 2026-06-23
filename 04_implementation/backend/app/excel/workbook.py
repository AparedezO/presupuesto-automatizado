from __future__ import annotations

import os
import tempfile
from copy import copy
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings
from app.excel.backups import create_workbook_backup
from app.excel.formulas import (
    FormulaIntegrityReport,
    collect_workbook_formulas,
    compare_formulas,
)


class WorkbookLockedError(RuntimeError):
    pass


class FormulaIntegrityError(RuntimeError):
    def __init__(self, report: FormulaIntegrityReport):
        super().__init__("Formula integrity check failed.")
        self.report = report


class ExcelWorkbook:
    def __init__(
        self,
        workbook_path: Path | None = None,
        backup_dir: Path | None = None,
    ) -> None:
        self.workbook_path = workbook_path or settings.workbook_path
        self.backup_dir = backup_dir or settings.backup_dir

    def load(self):
        return openpyxl.load_workbook(
            self.workbook_path,
            data_only=False,
            read_only=False,
            keep_links=True,
        )

    def assert_writable(self) -> None:
        try:
            with open(self.workbook_path, "r+b"):
                return
        except PermissionError as exc:
            raise WorkbookLockedError(
                f"Workbook is locked or open in another process: {self.workbook_path}"
            ) from exc

    def save_preserving_formulas(
        self,
        workbook,
        before_formulas: dict[str, str],
        allowed_added_formulas: Iterable[str] | None = None,
    ) -> tuple[Path, FormulaIntegrityReport]:
        after_formulas = collect_workbook_formulas(workbook)
        report = compare_formulas(
            before_formulas,
            after_formulas,
            allowed_added=set(allowed_added_formulas or []),
        )

        if report.changed or report.removed:
            raise FormulaIntegrityError(report)

        self.assert_writable()
        backup_path = create_workbook_backup(self.workbook_path, self.backup_dir)
        fd, temp_name = tempfile.mkstemp(
            suffix=self.workbook_path.suffix,
            dir=str(self.workbook_path.parent),
        )
        os.close(fd)
        temp_path = Path(temp_name)

        try:
            if hasattr(workbook, "calculation"):
                workbook.calculation.calcMode = "auto"
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True
            workbook.save(temp_path)
            openpyxl.load_workbook(temp_path, data_only=False, read_only=True).close()
            os.replace(temp_path, self.workbook_path)
        finally:
            if temp_path.exists():
                temp_path.unlink()

        return backup_path, report


def find_last_data_row(sheet: Worksheet, key_column: int, start_row: int) -> int:
    for row in range(sheet.max_row, start_row - 1, -1):
        if sheet.cell(row=row, column=key_column).value not in (None, ""):
            return row
    return start_row - 1


def copy_row_format_and_formulas(
    sheet: Worksheet,
    source_row: int,
    target_row: int,
    max_column: int | None = None,
) -> set[str]:
    max_column = max_column or sheet.max_column
    added_formulas: set[str] = set()
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)

        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)

        if isinstance(source.value, str) and source.value.startswith("="):
            try:
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value
            added_formulas.add(f"{sheet.title}!{target.coordinate}")
        else:
            target.value = None

    return added_formulas
