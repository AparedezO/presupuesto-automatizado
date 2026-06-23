from __future__ import annotations

from decimal import Decimal
from typing import Iterable

import openpyxl
from openpyxl.formula.translate import Translator

from app.excel.formulas import collect_workbook_formulas
from app.excel.workbook import (
    ExcelWorkbook,
    copy_row_format_and_formulas,
)
from app.schemas.apus import ApuCreate, ApuDetail, ApuDraftDetail, ApuDraftRead, ApuRead


APU_SHEET = "APU'S"
MATERIALS_SHEET = "L-MATERIALES"
HEADER_TEMPLATE_ROW = 3224
TITLE_TEMPLATE_ROW = 3225
DETAIL_TEMPLATE_ROWS = {
    "MATERIAL": 3226,
}
STANDARD_TEMPLATE_ROWS = (3231, 3232, 3233, 3234)
BLANK_TEMPLATE_ROW = 3235
STANDARD_ITEM_CODES = {"E&H", "GLV", "T", "H/C"}
MATERIAL_CODE_COL = 2


class ApuRepository:
    def __init__(self, excel: ExcelWorkbook | None = None) -> None:
        self.excel = excel or ExcelWorkbook()

    def _header_row_for_append(self, sheet) -> int:
        return sheet.max_row + 1

    def _find_apu_rows(self, sheet) -> list[int]:
        rows: list[int] = []
        for row in range(1, sheet.max_row + 1):
            marker = sheet.cell(row=row, column=7).value
            code = sheet.cell(row=row, column=2).value
            description = sheet.cell(row=row, column=3).value
            unit = sheet.cell(row=row, column=5).value

            if marker in {"COSTO ITEM", "COSTOITEM"} and code not in (None, ""):
                rows.append(row)
        return rows

    def list_apus(self) -> list[ApuRead]:
        workbook = self.excel.load()
        try:
            sheet = workbook[APU_SHEET]
            apus: list[ApuRead] = []

            for row in self._find_apu_rows(sheet):
                code = sheet.cell(row=row, column=2).value
                description = sheet.cell(row=row, column=3).value
                unit = sheet.cell(row=row, column=5).value

                apus.append(
                    ApuRead(
                        row_number=row,
                        codigo_apu=str(code).strip(),
                        descripcion_apu=str(description or "").strip(),
                        unidad=str(unit or "").strip(),
                        costo_total_formula=sheet.cell(row=row, column=8).value,
                    )
                )

            return apus
        finally:
            workbook.close()

    def search(self, term: str) -> list[ApuRead]:
        target = term.strip().upper()
        if not target:
            return self.list_apus()
        return [
            apu
            for apu in self.list_apus()
            if target in apu.codigo_apu.upper() or target in apu.descripcion_apu.upper()
        ]

    def exists(self, codigo_apu: str) -> bool:
        target = codigo_apu.strip().upper()
        return any(apu.codigo_apu.upper() == target for apu in self.list_apus())

    def _material_lookup(self, workbook) -> dict[str, dict[str, str]]:
        if MATERIALS_SHEET not in workbook.sheetnames:
            return {}

        sheet = workbook[MATERIALS_SHEET]
        lookup: dict[str, dict[str, str]] = {}
        for row in range(1, sheet.max_row + 1):
            code = sheet.cell(row=row, column=2).value
            if code in (None, ""):
                continue

            normalized_code = str(code).strip()
            lookup[normalized_code.upper()] = {
                "descripcion": str(sheet.cell(row=row, column=3).value or "").strip(),
                "und": str(sheet.cell(row=row, column=5).value or "").strip(),
                "costo": str(sheet.cell(row=row, column=7).value or ""),
            }
        return lookup

    def get_draft_from_row(self, row_number: int) -> ApuDraftRead:
        workbook = self.excel.load()
        value_workbook = openpyxl.load_workbook(
            self.excel.workbook_path,
            data_only=True,
            read_only=False,
            keep_links=True,
        )
        try:
            sheet = workbook[APU_SHEET]
            value_sheet = value_workbook[APU_SHEET]
            if row_number < 1 or row_number > sheet.max_row:
                raise ValueError(f"APU row does not exist: {row_number}")

            marker = sheet.cell(row=row_number, column=7).value
            code = sheet.cell(row=row_number, column=2).value
            if marker not in {"COSTO ITEM", "COSTOITEM"} or code in (None, ""):
                raise ValueError(f"Row is not an APU header: {row_number}")

            material_lookup = self._material_lookup(workbook)
            details: list[ApuDraftDetail] = []

            for row in range(row_number + 2, sheet.max_row + 1):
                item_code = sheet.cell(row=row, column=MATERIAL_CODE_COL).value
                if item_code in (None, ""):
                    break

                normalized_code = str(item_code).strip()
                if normalized_code.upper() in STANDARD_ITEM_CODES:
                    break

                material = material_lookup.get(normalized_code.upper(), {})
                quantity_value = sheet.cell(row=row, column=6).value
                quantity_formula = (
                    quantity_value.strip()
                    if isinstance(quantity_value, str) and quantity_value.strip().startswith("=")
                    else None
                )
                quantity_display = (
                    value_sheet.cell(row=row, column=6).value
                    if quantity_formula
                    else quantity_value
                )
                if quantity_display in (None, ""):
                    quantity_display = 0 if quantity_formula else 1

                details.append(
                    ApuDraftDetail(
                        row_number=row,
                        source_row_number=row,
                        tipo_item="MATERIAL",
                        codigo_item=normalized_code,
                        cantidad=str(quantity_display),
                        cantidad_formula=quantity_formula,
                        descripcion=material.get("descripcion")
                        or str(sheet.cell(row=row, column=3).value or "").strip(),
                        und=material.get("und")
                        or str(sheet.cell(row=row, column=5).value or "").strip(),
                        costo=material.get("costo")
                        or str(sheet.cell(row=row, column=7).value or ""),
                    )
                )

            if not details:
                raise ValueError(f"APU has no material rows: {code}")

            return ApuDraftRead(
                base_row_number=row_number,
                base_codigo_apu=str(code).strip(),
                codigo_apu="",
                descripcion_apu=str(sheet.cell(row=row_number, column=3).value or "").strip(),
                unidad=str(sheet.cell(row=row_number, column=5).value or "UN").strip(),
                detalles=details,
            )
        finally:
            workbook.close()
            value_workbook.close()

    def _apply_header(self, sheet, row: int, payload: ApuCreate) -> set[str]:
        allowed = copy_row_format_and_formulas(sheet, HEADER_TEMPLATE_ROW, row)
        sheet.cell(row=row, column=2).value = payload.codigo_apu.strip()
        sheet.cell(row=row, column=3).value = payload.descripcion_apu.strip()
        sheet.cell(row=row, column=4).value = "MARCA"
        sheet.cell(row=row, column=5).value = payload.unidad.strip()
        sheet.cell(row=row, column=6).value = None
        sheet.cell(row=row, column=7).value = "COSTO ITEM"
        return allowed

    def _apply_title_row(self, sheet, row: int) -> set[str]:
        allowed = copy_row_format_and_formulas(sheet, TITLE_TEMPLATE_ROW, row)
        sheet.cell(row=row, column=2).value = "CODIGO"
        sheet.cell(row=row, column=3).value = "DETALLE"
        sheet.cell(row=row, column=4).value = None
        sheet.cell(row=row, column=5).value = "UNIDAD"
        sheet.cell(row=row, column=6).value = "CANTIDAD"
        sheet.cell(row=row, column=7).value = "V/UNITARIO"
        sheet.cell(row=row, column=8).value = "V/PARCIAL"
        sheet.cell(row=row, column=9).value = "M-E&H"
        return allowed

    def _apply_detail_row(self, sheet, row: int, detail: ApuDetail) -> set[str]:
        if detail.tipo_item != "MATERIAL":
            raise ValueError("Los APUs solo permiten agregar materiales manualmente.")
        template_row = DETAIL_TEMPLATE_ROWS[detail.tipo_item]
        allowed = copy_row_format_and_formulas(sheet, template_row, row)
        sheet.cell(row=row, column=2).value = detail.codigo_item.strip()

        quantity = detail.cantidad
        quantity_formula = detail.cantidad_formula
        if not quantity_formula and isinstance(quantity, str) and quantity.strip().startswith("="):
            quantity_formula = quantity.strip()

        if quantity_formula:
            if not detail.source_row_number:
                raise ValueError(
                    f"La cantidad formula del material {detail.codigo_item} no tiene fila base."
                )
            try:
                sheet.cell(row=row, column=6).value = Translator(
                    quantity_formula,
                    origin=f"F{detail.source_row_number}",
                ).translate_formula(f"F{row}")
            except Exception:
                sheet.cell(row=row, column=6).value = quantity_formula
            allowed.add(f"{sheet.title}!F{row}")
        else:
            try:
                sheet.cell(row=row, column=6).value = float(
                    Decimal(str(quantity).replace(",", "."))
                )
            except Exception as exc:
                raise ValueError(
                    f"La cantidad del material {detail.codigo_item} debe ser numerica o formula."
                ) from exc
        return allowed

    def _apply_standard_row(self, sheet, source_row: int, target_row: int) -> set[str]:
        allowed = copy_row_format_and_formulas(sheet, source_row, target_row)
        for column in range(1, sheet.max_column + 1):
            source_value = sheet.cell(row=source_row, column=column).value
            if not (isinstance(source_value, str) and source_value.startswith("=")):
                sheet.cell(row=target_row, column=column).value = source_value
        return allowed

    def _apply_blank_row(self, sheet, row: int) -> set[str]:
        allowed = copy_row_format_and_formulas(sheet, BLANK_TEMPLATE_ROW, row)
        for column in range(2, sheet.max_column + 1):
            sheet.cell(row=row, column=column).value = None
        return allowed

    def _apply_apu_code_formulas(self, sheet, start_row: int, blank_row: int) -> set[str]:
        allowed: set[str] = set()
        sheet.cell(row=start_row, column=1).value = (
            f'=+IF(B{blank_row}="",B{start_row},A{blank_row})'
        )
        allowed.add(f"{sheet.title}!A{start_row}")

        for row in range(start_row + 1, blank_row + 1):
            sheet.cell(row=row, column=1).value = f'=+IF(B{row - 1}="",B{row},A{row - 1})'
            allowed.add(f"{sheet.title}!A{row}")

        return allowed

    def create_apu(self, payload: ApuCreate) -> dict:
        if not payload.detalles:
            raise ValueError("An APU requires at least one detail row.")
        if any(detail.tipo_item != "MATERIAL" for detail in payload.detalles):
            raise ValueError("Los APUs solo permiten agregar materiales manualmente.")
        if self.exists(payload.codigo_apu):
            raise ValueError(f"APU code already exists: {payload.codigo_apu}")

        workbook = self.excel.load()
        try:
            before_formulas = collect_workbook_formulas(workbook)
            sheet = workbook[APU_SHEET]
            start_row = self._header_row_for_append(sheet)
            detail_start = start_row + 2
            detail_end = detail_start + len(payload.detalles) - 1
            standard_start = detail_end + 1
            standard_end = standard_start + len(STANDARD_TEMPLATE_ROWS) - 1
            blank_row = standard_end + 1

            allowed_added_formulas: set[str] = set()
            allowed_added_formulas.update(self._apply_header(sheet, start_row, payload))
            allowed_added_formulas.update(self._apply_title_row(sheet, start_row + 1))

            for index, detail in enumerate(payload.detalles):
                target_row = detail_start + index
                allowed_added_formulas.update(self._apply_detail_row(sheet, target_row, detail))

            for offset, source_row in enumerate(STANDARD_TEMPLATE_ROWS):
                target_row = standard_start + offset
                allowed_added_formulas.update(
                    self._apply_standard_row(sheet, source_row, target_row)
                )

            allowed_added_formulas.update(self._apply_blank_row(sheet, blank_row))
            allowed_added_formulas.update(
                self._apply_apu_code_formulas(sheet, start_row, blank_row)
            )

            sheet.cell(row=start_row, column=8).value = f"=SUM(H{detail_start}:H{standard_end})"
            sheet.cell(row=start_row, column=9).value = f"=+SUM(H{standard_start}:H{standard_end})"

            backup_path, report = self.excel.save_preserving_formulas(
                workbook,
                before_formulas,
                allowed_added_formulas=allowed_added_formulas,
            )

            return {
                "header_row": start_row,
                "detail_rows": len(payload.detalles) + len(STANDARD_TEMPLATE_ROWS) + 1,
                "backup_path": str(backup_path),
                "formula_integrity_ok": report.ok,
            }
        finally:
            workbook.close()
